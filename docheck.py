import json
import shutil
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
from openpyxl import load_workbook
import _thread
import tkinter
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from webdriver_manager.core.os_manager import OperationSystemManager
from tkinter.messagebox import showinfo
from openpyxl.styles import PatternFill
import oprate_redmine

waiverList = {
    "libcore.java.net.InetAddressTest#test_getByName_invalid[1]": "ipv6",
    "libcore.java.net.InetAddressTest#test_isReachable_by_ICMP": "ipv6",
    "libcore.java.net.SocketTest#testSocketTestAllAddresses": "ipv6",
    "android.net.cts.ConnectivityManagerTest#testOpenConnection": "foreign SIM card",
    "android.net.cts.MultinetworkApiTest#testNativeDatagramTransmission": "ipv6",
    "android.net.cts.DnsTest#testDnsWorks": "ipv6",
    "com.google.android.location.gts.gnss.GnssPseudorangeVerificationTest#testPseudoPosition": "weak GNSS signal indoor",
    "com.google.android.location.gts.gnss.GnssPseudorangeVerificationTest#testPseudorangeValue": "weak GNSS signal indoor",
    "android.location.cts.GnssPseudorangeVerificationTest#testPseudoPosition": "weak GNSS signal indoor",
    "android.location.cts.GnssPseudorangeVerificationTest#testPseudorangeValue": "weak GNSS signal indoor",
    "android.location.cts.gnss.GnssPseudorangeVerificationTest#testPseudoPosition": "weak GNSS signal indoor",
    "com.google.android.media.gts.WidevineDashPolicyTests#testL3OfflineCannotPersist": "waiver",
}


# path：报告路径，必须是下一级包含"cts","vts"等文件夹的目录，也即一版软件的报告路径
# 返回所有test_result_failure报告的路径的列表
def get_report(path):
    list_dir = os.listdir(path)
    path_report = []
    for i in list_dir:
        if i == "cts" or i == "vts" or i == "sts" or i == "gts" or i == "gsi" or i == "cts-instant":
            sub_list = os.path.join(path, i)
            try:
                for j in os.listdir(sub_list):
                    if j[:3] == "202":
                        sub_list2 = os.path.join(sub_list, j)
                        try:
                            for q in os.listdir(sub_list2):
                                if q == "test_result_failures_suite.html" or q == "test_result_failures.html":
                                    path_report.append(os.path.join(sub_list2, q))
                        except Exception as e:
                            print("非目录，跳过")
            except Exception as e:
                print("跳过")
    return path_report


def fill_color(workbook_DL):
    for x in ["CTS", "GTS", "CTS-ON-GSI", "VTS", "STS"]:
        i = 4  # 根据模板，从第四行开始填充失败项的颜色
        while True:
            if not workbook_DL[x]["A" + str(i)].value is None:
                workbook_DL[x]["A" + str(i)].fill = PatternFill(start_color="ffff00", fill_type="solid")
                if workbook_DL[x]["B" + str(i)].value in waiverList.keys():
                    workbook_DL[x]["E" + str(i)] = waiverList[workbook_DL[x]["B" + str(i)].value]
                    for j in ["B", "C", "D", "E"]:
                        workbook_DL[x][j + str(i)].fill = PatternFill(start_color="92d050", fill_type="solid")
                else:
                    for j in ["B", "C", "D", "E"]:
                        workbook_DL[x][j + str(i)].fill = PatternFill(start_color="ff0000", fill_type="solid")
                i = i + 1
            else:
                break
        workbook_DL[x].append(["Incomplete Modules"])
        workbook_DL[x]["A" + str(i + 1)].fill = PatternFill(start_color="a5c639", fill_type="solid")

    # report:报告文件的路径
    # 返回该报告中的一些信息


def getinfo(report, browser):
    # print("broser you choose:" + browser.get())
    infodict = dict()
    driver = None
    operationSystemManager = OperationSystemManager()
    with open("webview_version.json", "r") as f:
        f_j = json.load(f)
        web_version_edge_json = f_j["edge_version"]
        # print(f"web_version_edge:{web_version_edge_json}")
        web_version_chrome_json = f_j["chrome_version"]
        # print(f"web_version_chrome:{web_version_chrome_json}")

    try:
        if browser.get() == "chrome":
            chrome_version = operationSystemManager.get_browser_version_from_os("google-chrome")
            print("local chrome version" + chrome_version)
            # 设置selenium使用chrome的无头模式
            chrome_options = Options()
            chrome_options.add_argument('headless')  # 设置option,隐藏浏览器界面
            if web_version_chrome_json == chrome_version and "chromedriver.exe" in os.listdir(os.getcwd()):
                driver = webdriver.Chrome(f"{os.getcwd()}/chromedriver.exe", options=chrome_options)
            else:
                # 在启动浏览器时加入配置
                driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)  # 自动获取chrome浏览器的驱动
                shutil.copy(ChromeDriverManager().install(), os.getcwd())
                with open("webview_version.json", "w") as f:
                    f_j["chrome_version"] = chrome_version
                    # json.dump(f_j, f)
                    f.write(json.dumps(f_j))
        elif browser.get() == "edge":
            edge_version = operationSystemManager.get_browser_version_from_os("edge")
            print(edge_version)
            options = {
                # "browserName": "MicrosoftEdge",
                # "version": "",
                # "platform": "WINDOWS",
                "ms:edgeOptions": {
                    "extensions": [], "args": ["--headless"]  # 添加隐藏浏览器页面运作参数
                }
            }

            if web_version_edge_json == edge_version and "msedgedriver.exe" in os.listdir(os.getcwd()):
                driver = webdriver.Edge(f"{os.getcwd()}/msedgedriver.exe", capabilities=options)
            else:
                driver = webdriver.Edge(EdgeChromiumDriverManager().install(), capabilities=options)  # 自动获取edge浏览器的驱动
                shutil.copy(EdgeChromiumDriverManager().install(), os.getcwd())
                with open("webview_version.json", "w") as f:
                    f_j["edge_version"] = edge_version
                    json.dump(f_j, f)
                    # f.write(json.dumps(f_j))

    except Exception as e:
        print(e)
        if str(e).startswith("HTTPSConnectionPool"):
            print("获取浏览器驱动失败，请连接网络后重试!")
    driver.get('file:///' + report)
    # 等待加载，最多等待20秒
    driver.implicitly_wait(20)
    # browser.maximize_window()  # 窗口最大化

    li = driver.find_elements_by_xpath("//td[@class='rowtitle']/../td[2]")
    infodict["suite_plan"] = li[0].text
    infodict["suite_build"] = li[1].text
    infodict["host_info"] = li[2].text
    infodict["time_info"] = li[3].text
    infodict["case_pass"] = li[4].text
    infodict["case_fail"] = li[5].text
    infodict["modules_done"] = li[6].text
    infodict["modules_total"] = li[7].text
    infodict["finger_print"] = li[8].text
    infodict["security_patch"] = li[9].text
    infodict["release_sdk"] = li[10].text
    infodict["ABIs"] = li[11].text

    details = driver.find_elements_by_class_name("testdetails")
    all_fail = []  # 所有fail项的列表
    for fail_module in details:
        module_name = fail_module.find_element_by_class_name("module").text
        fails = fail_module.find_elements_by_class_name("testname")
        for fail_item in fails:
            fail = {"module": module_name, "name": fail_item.text,
                    "detail": "null for temp"}  # 每个fail为字典，module：所属模块，name：失败项case名称，detail：报错信息
            all_fail.append(fail)
    infodict["fails"] = all_fail
    # print(all_fail)
    driver.quit()
    return infodict


class AutoWork:
    entry1 = None
    ifDL = None
    main_window = None
    browser_choose = None

    # 在表格模板中填充信息
    # all_info：所有报告信息字典的列表
    # path：汇总表格文件的输出目录
    def write_xl(self, all_info, path=""):
        workbook_DL = load_workbook(filename="DL_xTS_Test_Report.xlsx")
        sheet_DL_Summary = workbook_DL["Summary"]
        sheet_DL_CTS = workbook_DL["CTS"]
        sheet_DL_GTS = workbook_DL["GTS"]
        sheet_DL_VTS = workbook_DL["VTS"]
        sheet_DL_CTS_ON_GSI = workbook_DL["CTS-ON-GSI"]
        sheet_DL_STS = workbook_DL["STS"]
        sheet_DL_CTSV = workbook_DL["CTS_VERIFIER"]

        workbook = load_workbook(filename="case.xlsx")
        sheet1 = workbook["case汇总"]
        sheet2 = workbook["失败项汇总"]
        row0 = ["plan", "tool", "case_all_test", "case_pass", "case_fail", "modules", "finger_print"]
        sheet1.append(row0)
        data = []
        for info in all_info:
            plan = info["suite_plan"]
            build = info["suite_build"]
            case_pass = info["case_pass"]
            case_fail = info["case_fail"]
            modules_done = info["modules_done"]
            modules_total = info["modules_total"]
            security_patch = info["security_patch"]
            finger_print = info["finger_print"]
            fails = info["fails"]  # info["fails"]是个列表,其中每个fail元素是字典

            case_all_test = int(case_pass) + int(case_fail)
            row = [plan, build, case_all_test, int(case_pass), int(case_fail), modules_done + "/" + modules_total,
                   finger_print]
            data.append(row)  # 将信息直接附在后面
            for fail in fails:
                row_fail = [plan, fail["module"], fail["name"]]  # , fail["detail"]
                sheet2.append(row_fail)

            # 自动填充工具信息及模块和case数到模板中的固定位置，同一plan多个报告取total_case数量最多的
            p = plan.split('/')
            s = build.split('/')
            tool = p[0] + s[0]
            if plan == "CTS / cts" or plan == "CTS / cts-retry":
                sheet1['B1'] = tool
                sheet1['C1'] = tool
                sheet1['D1'] = tool
                if sheet1["C3"].value is None or int(modules_total) > sheet1["C3"].value:
                    sheet1["C3"] = int(modules_total)
                if sheet1["C4"].value is None or int(modules_total) > sheet1["C4"].value:
                    sheet1["C4"] = case_all_test
                if self.ifDL.get() == "DL":
                    sheet_DL_Summary['B2'] = finger_print
                    sheet_DL_Summary['B3'] = security_patch
                    sheet_DL_Summary['C6'] = build
                    sheet_DL_Summary['C10'] = build
                    sheet_DL_Summary['D6'] = modules_done + "/" + modules_total
                    sheet_DL_Summary['F6'] = int(case_fail)
                    for fail in fails:
                        row_fail = [fail["module"], fail["name"]]  # , fail["detail"]
                        sheet_DL_CTS.append(row_fail)

            if plan == "VTS / cts-on-gsi" or plan == "VTS / cts-on-gsi-retry":
                # sheet1['D1'] = tool
                if sheet1["D3"].value is None or int(modules_total) > sheet1["D3"].value:
                    sheet1["D3"] = int(modules_total)
                if sheet1["D4"].value is None or int(modules_total) > sheet1["D4"].value:
                    sheet1["D4"] = case_all_test
                if self.ifDL.get() == "DL":
                    sheet_DL_Summary['C9'] = build
                    sheet_DL_Summary['D9'] = modules_done + "/" + modules_total
                    sheet_DL_Summary['F9'] = int(case_fail)
                    for fail in fails:
                        row_fail = [fail["module"], fail["name"]]  # , fail["detail"]
                        sheet_DL_CTS_ON_GSI.append(row_fail)

            if plan == "VTS / vts":
                sheet1['E1'] = tool
                if sheet1["E3"].value is None or int(modules_total) > sheet1["E3"].value:
                    sheet1["E3"] = int(modules_total)
                if sheet1["E4"].value is None or int(modules_total) > sheet1["E4"].value:
                    sheet1["E4"] = case_all_test
                if self.ifDL.get() == "DL":
                    sheet_DL_Summary['C8'] = build
                    sheet_DL_Summary['D8'] = modules_done + "/" + modules_total
                    sheet_DL_Summary['F8'] = int(case_fail)
                    for fail in fails:
                        row_fail = [fail["module"], fail["name"]]  # , fail["detail"]
                        sheet_DL_VTS.append(row_fail)

            if plan == "GTS / gts":
                sheet1['F1'] = tool
                if sheet1["F3"].value is None or int(modules_total) > sheet1["F3"].value:
                    sheet1["F3"] = int(modules_total)
                if sheet1["F4"].value is None or int(modules_total) > sheet1["F4"].value:
                    sheet1["F4"] = case_all_test
                if self.ifDL.get() == "DL":
                    sheet_DL_Summary['C7'] = build
                    sheet_DL_Summary['D7'] = modules_done + "/" + modules_total
                    sheet_DL_Summary['F7'] = int(case_fail)
                    for fail in fails:
                        row_fail = [fail["module"], fail["name"]]  # , fail["detail"]
                        sheet_DL_GTS.append(row_fail)

            if plan == "STS / sts-engbuild" or plan == "STS / sts-dynamic-incremental" or \
                    plan == "STS / sts-dynamic-full":
                sheet1['G1'] = tool
                if sheet1["G3"].value is None or int(modules_total) > sheet1["G3"].value:
                    sheet1["G3"] = int(modules_total)
                if sheet1["G4"].value is None or int(modules_total) > sheet1["G4"].value:
                    sheet1["G4"] = case_all_test
                if self.ifDL.get() == "DL":
                    sheet_DL_Summary['C5'] = build
                    sheet_DL_Summary['D5'] = modules_done + "/" + modules_total
                    sheet_DL_Summary['F5'] = int(case_fail)
                    for fail in fails:
                        row_fail = [fail["module"], fail["name"]]  # , fail["detail"]
                        sheet_DL_STS.append(row_fail)

        for r in data:
            sheet1.append(r)
        workbook.save(filename=path + r"\汇总.xlsx")
        if self.ifDL.get() == "DL":
            fill_color(workbook_DL)
            workbook_DL.save(filename=path + r"\DL_xTS_Test_Report.xlsx")

    def do_my_print(self, path):
        _thread.start_new_thread(self.real_do, (path,))

    def real_do(self, path):
        label2 = tkinter.Label(self.main_window, text="执行中...")
        label2.pack()
        print("path: " + path)
        list_dir = os.listdir(path)
        sub = False
        for i in list_dir:
            if i == "CN" or i == "EU" or i == "RU" or i == "US" or i == "WWAN" or i == "WLAN":
                sub = True
                path_sub = os.path.join(path, i)
                print("path_sub:" + path_sub)
                self.real_real_do(path_sub)
        if not sub:
            self.real_real_do(path)
        showinfo(title="完成", message="完成！汇总表格已保存：\n" + path)
        label2.pack_forget()

    def real_real_do(self, path):
        all_info = []  # 所有报告的信息字典的列表
        for i in get_report(path):
            infodict = getinfo(i, self.browser_choose)
            all_info.append(infodict)
            print("dict:" + str(infodict))
        self.write_xl(all_info, path)
        oprate_redmine.new_all_bugs(all_info)
        print("完成！" + path)

    def init_window(self):
        self.main_window = tkinter.Tk()
        self.main_window.title("Auto Check")
        self.main_window.geometry("300x200")
        label1 = tkinter.Label(self.main_window, text="请输入路径：")
        label1.pack()
        entry1 = tkinter.Entry(self.main_window)
        entry1.pack()
        self.browser_choose = tkinter.StringVar(value="edge")
        radioBtn_browser_edge = tkinter.Radiobutton(self.main_window, text="Microsoft Edge",
                                                    variable=self.browser_choose, value="edge")
        radioBtn_browser_chrome = tkinter.Radiobutton(self.main_window, text="Chrome", variable=self.browser_choose,
                                                      value="chrome")
        radioBtn_browser_chrome.pack()
        radioBtn_browser_edge.pack()
        button1 = tkinter.Button(self.main_window, text="开始", command=lambda: self.do_my_print(entry1.get()))
        button1.pack()
        self.ifDL = tkinter.StringVar(value="DL")
        radioBtnA = tkinter.Radiobutton(self.main_window, text="使用DL报告模板", variable=self.ifDL, value="DL")
        radioBtnA.pack()
        # radioBtnB = tkinter.Radiobutton(self.main_window, text="使用非DL报告模板", variable=self.ifDL, value="No-DL")
        # radioBtnB.pack()

        # button_test = tkinter.Button(self.main_window, text="测试按钮", command=lambda: print(self.ifDL.get()))
        # button_test.pack()

        self.main_window.mainloop()


if __name__ == '__main__':
    auto = AutoWork()
    auto.init_window()
